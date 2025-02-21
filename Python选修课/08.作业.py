# 01 写入文件
'''
with open(r"f_write.txt", "a+") as f:
    while(1):
        char = input("请输入字符：\n")
        if char == '@':
            break
        else:
            f.write(char)

# 02 存到excel
from faker import Faker
#import openpyxl as pxl
from xlwt import Workbook

fake = Faker('zh_CN')

w = Workbook()
ws = w.add_sheet('faker_info')

headers = ["姓名", "性别", "年龄", "邮箱", "手机"]

for col, header in enumerate(headers):
    ws.write(0, col, header)

for row in range(1, 11):
    name = fake.name()
    sex = fake.random_element(["男", "女"])
    age = fake.random_int(min=18, max=80)
    for col, info in enumerate(infos):
        ws.write(row, col, info)

w.save('faker_info.xls')

# 03 flourish
import csv
from openpyxl import load_workbook
# import pandas as pd

# 写入xlsx
wb = load_workbook('flourish.xlsx')
ws = wb.active
ws.cell(1, 1, 'Country/Region')

medal_counts ={}

with open(r"D:\code\Python\Python选修课\Athletes_summer_games.csv", 'r', encoding='utf-8') as f:
    csv_reader = csv.reader(f)

    for row in csv_reader:
        noc_value = row['NOC']
        medal_value = row['Medal']
        year_value = row['Year']


        if noc_value not in medal_counts:
            medal_counts[noc_value] = {}
        if year_value not in medal_counts[noc_value]:
            medal_counts[noc_value][year_value] = 0

        if medal_value == 'Gold':
            medal_counts[noc_value][year_value] += 1

country_header = 'Country/Region'
for noc_value, year_values in medal_counts.items():
    country_index = ws['Country/Region'].index(noc_value) + 2
    for year_value, count in year_values.items():
        year_column = int(year_value) - 1895
        ws.cell(country_index, year_column, count)

wb.save(r'flourish.xlsx')
'''

import csv
import openpyxl as pxl

with open(r"D:\code\Python\Python选修课\Athletes_summer_games.csv", 'r', encoding='utf-8') as f:
    csv_reader = csv.reader(f)
    raw_data = []
    for each in csv_reader:
        raw_data.append(each)

header =raw_data[0]
raw_data = raw_data[1:len(raw_data)]



# 包含多少年
years = []
for record in raw_data:
    years.append(record[7])
years = sorted(set(years))

# 包含多少个国家/地区
countries = []
for record in raw_data:
    countries.append(record[5])
countries = sorted(set(countries))

medal_counts = {}
for year in years:
    medal_counts[year] = {}
    for country in countries:
        medal_counts[year][country] = 0

for record in raw_data:
    year = record[7]
    country = record[5]
    if record[-1] == 'Gold':
        medal_counts[year][country] += 1

#写入到xlsx文件中
wb = pxl.Workbook()
ws = wb.active
ws.cell(1, 1, 'Country/Region')
for i in range(len(years)):
    year = years[i]
    ws.cell(1, i+2, year)

for i in range(len(countries)):
    country = countries[i]
    ws.cell(i+2, 1, country)

for i in range(len(countries)):
    country = countries[i]
    for j in range(len(years)):
        year = years[j]
        ws.cell(i+2, j+2, medal_counts[year][country])

wb.save(r'flourish.xlsx')
wb.close()
'''  
# 去除错误数据和空数据
exist_wrong_record = True
while (exist_wrong_record == True):
    exist_wrong_record = False
    for record in raw_data:
        if (('N/A') in record) or ('' in record):
            exist_wrong_record = True
            break
    idx = raw_data.index(record)
    del raw_data[idx]

# 原始数据包含多少年
years = []
for record in raw_data:
    years.append(record[0])
years = sorted(set(years))

# 原始数据包含多少个国家/地区
countries = []
for record in raw_data:
    countries.append(record[2])
countries = sorted(set(countries))


for i in range(len(years)):
    ws.cell(1, i+2, years[i])

for i in range(len(countries)):
    ws.cell(i+2, 1, countries[i])
'''










